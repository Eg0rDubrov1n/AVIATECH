import json

import archive
import os

import pymysql
from aiogram import Bot
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, CallbackQuery

from core.keyboards.reply import KeyGlobal
from core.setings import settings
from core.unit.SignalState import Form, Specialist, FormMesegeInlineKeyboard

import aspose.zip as az
from openpyxl import Workbook

from core.keyboards.inline import generatorWorkerKeyBoard, generatorMainKeyBoard, FORMPRINT


# class Signal(StatesGroup):
#     name_TasksSignal = 0
#     name_of_the_specialist = 0
#     description = 0
#     download_zip = 0
#     MenuInChat = False

# def SbrosSignal():
#     Signal.name_TasksSignal = 0
#     Signal.name_of_the_specialist = 0
#     Signal.description = 0
#     Signal.download_zip = 0

async def newProject(message: Message, state: FSMContext) -> None:
    # print(state)
    # await bot.delete_message(message.chat.id, message_id=message.message_id)
    # try:
    #     # Все сообщения, начиная с текущего и до первого (message_id = 0)
    #     for i in range(message.message_id, 0, -1):
    #         await bot.delete_message(message.from_user.id, i)
    # except TelegramBadRequest as ex:
    #     # Если сообщение не найдено (уже удалено или не существует),
    #     # код ошибки будет "Bad Request: message to delete not found"
    #     if ex.message == "Bad Request: message to delete not found":
    #         print("Все сообщения удалены")

    # Signal.MenuInChat = True
    await state.clear()
    data = await state.get_data()
    print(data.get("name_Tasks"))
    print(("NEWWWWWWWWWWWWWWWW"))
    await FORMPRINT(state)

    try:
        await message.answer(text=f"Здравствуйте",
            reply_markup=await generatorMainKeyBoard(state)
        )
    except Exception:
        print("НИХЕРА НЕ ИЗМЕНИЛОСЬ",Exception)
    FormMesegeInlineKeyboard.MesegeID =  str(int(message.message_id) + 1)
    FormMesegeInlineKeyboard.ChatID =  message.chat.id

async def name_pjoject(call: CallbackQuery, state: FSMContext):
    print("name_pjoject")
    await state.set_state(Form.name_Tasks)
    data = await state.get_data()
    await FORMPRINT(state)
    await call.answer("Введите название проекта")

async def name_of_the_specialist(call: CallbackQuery, state: FSMContext):
    await state.set_state(Form.designated_People)
    print("name_of_the_specialist")
    await call.answer("Выберите специалиста")
    await call.message.edit_reply_markup(
        reply_markup=await generatorWorkerKeyBoard(state)
    )

async def description(call: CallbackQuery, state: FSMContext):
    print("description")
    await state.set_state(Form.Description)
    await call.answer("Введите краткое описание проекта")

async def download_zip(call: CallbackQuery, state: FSMContext):
    print("download_zip")
    await call.answer("Загрузите Zip файл")
    await state.set_state(Form.download_zip)



def DelDir(directory):
    # Используйте os.walk для обхода дерева каталогов
    for root, dirs, files in os.walk(directory):
        # Для каждого файла в каталоге
        for file in files:
            # Постройте полный путь к файлу
            file_path = os.path.join(root, file)
            # Удалите файл
            os.remove(file_path)
        # Для каждого подкаталога в директории
        for dir in dirs:
            # Постройте полный путь к подкаталогу
            dir_path = os.path.join(root, dir)
            # Удалите подкаталог
            os.rmdir(dir_path)
    # Удалите каталог верхнего уровня
    os.rmdir(directory)

def CreateNewExeleFile(data,name):
    wb = Workbook()
    ws = wb.active
    Column = 1
    for key, item in data.items():
        print(key, item)
        if item != None and key != 'download_zip':
            ws.cell(row=Column, column=1, value=key)
            ws.cell(row=Column, column=2, value=item)
            Column += 1
    wb.save(f"{name}/test.xlsx")

def WriteInSQL(data):
    connect = pymysql.connect(
        host="127.0.0.1",
        port=3306,
        user="root",
        password="root",
        database="test12",
        cursorclass=pymysql.cursors.DictCursor
    )
    isstr_queryKey = list()
    isstr_queryValues = list()
    for key, item in data.items():
        if item != None and key != 'download_zip':
            isstr_queryKey.append(f"'{key}'")
            if key == "designated_People":
                isstr_queryValues.append(f"JSON_ARRAY({item})".replace('[','').replace(']',''))
            else:
                isstr_queryValues.append(f"'{item}'")
    isstr_queryKey = ','.join(isstr_queryKey).replace('\'', '`')
    isstr_queryValues = ','.join(isstr_queryValues)
    with connect.cursor() as cursor:
        isstr_query = (f"INSERT INTO `tasks` ({isstr_queryKey}) VALUES ({isstr_queryValues});")
        cursor.execute(isstr_query)
        connect.commit()
    print("____________________________________________")
    #---------------------------------------------------------------------
    #---------------------------------------------------------------------
    with connect.cursor() as cursor:
        idRequest = "SELECT AUTO_INCREMENT FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'group';"
        cursor.execute(idRequest)
        idTask = cursor.fetchall()[0]["AUTO_INCREMENT"]
        for idPeople in data.get("designated_People"):
            sqlCommand = f"SELECT  IDTasks FROM `group` WHERE ID={idPeople}"
            print(f"sqlCommand------------------------------------------{sqlCommand}")
            cursor.execute(sqlCommand)
            arr = cursor.fetchall()[0]['IDTasks']
            print("arr------------------------------------------")
            print(arr)
            if arr == None:
                arr = list()
            arr.append(str(idTask - 1))
            sqlCommand = f"UPDATE `group` SET IDTasks = {f'JSON_ARRAY({arr})'.replace('[','').replace(']','')} WHERE ID={idPeople};"
            cursor.execute(sqlCommand)
            connect.commit()







async def send(call: CallbackQuery, state: FSMContext, bot : Bot):

    data = await state.get_data()

    # data = await state.get_data()
    # print(f"-----------------------{state}--------------------")
    print(data.get("designated_People"))
    if data.get("name_Tasks") != None:
        # print("Finish--->")
        # print(data.get("download_zip"))
        name = settings.bots.path_save + data.get("name_Tasks")
        # fileTEXT = open(f"{name}/{data.get('name_Tasks')}.txt", 'w+')
        if data.get("download_zip") != None:
            os.mkdir(name)
            file_id = data.get('download_zip')
            print(f"-----------------------{file_id}----------------------")
            file = await bot.get_file(file_id)
            file_path = file.file_path
            print(f"{name}text.{file_path[file_path.rfind('.'):]}")
            await bot.download_file(file_path, f"{name}/text.{file_path[file_path.rfind('.') + 1:]}")
            with az.Archive() as archive:
                # Добавить папку в zip
                archive.create_entries(name)
                archive.save(f"{name}.zip")
            DelDir(name)
        # # print(f"-----------------------{file_path}----------------------")
        # # await bot.download_file(data.get('name_Tasks'), "text.txt")
        # for key, item in data.items():
        #     print(key, item)
        #     if item != None and key != 'download_zip':
        #         try:
        #             fileTEXT.write(f"{key}: {item} \n")
        #         except Exception:
        #             print("ERROR")
        #             # print(Exception)
        # CreateNewExeleFile(data,name)
        # fileTEXT.close()
        WriteInSQL(data)

    else:
        await call.message.answer(text="Error:Неуказанно имя!!!")
async def exit(call: CallbackQuery, state: FSMContext):
    await state.clear()
    # SbrosSignal()
    await call.message.edit_reply_markup()
    await call.message.answer("What Project?",reply_markup=KeyGlobal)

async def exitMainKey(call: CallbackQuery, state: FSMContext):
    await call.message.edit_reply_markup("What Project?", reply_markup=await generatorMainKeyBoard(state))



async def name_of_the_specialistPoint2(call: CallbackQuery, state: FSMContext,bot:Bot):
    data = await state.get_data()
    temporaryArray = list()
    if data.get("designated_People") != None:
        temporaryArray = data.get("designated_People")
    if call.data in temporaryArray:
        temporaryArray.remove(call.data)
    else:
        temporaryArray.append(call.data)

    print(temporaryArray)
    await state.update_data(designated_People=temporaryArray)
    print("State UPDETE designated_People")

    await call.message.edit_reply_markup(
        reply_markup=await generatorWorkerKeyBoard(state)
    )

    print(call.data)
    # await bot.edit_message_text(text="ОБНОВА", reply_markup=await generatorWorkerKeyBoard(state),
    #                             chat_id=FormMesegeInlineKeyboard.ChatID, message_id=FormMesegeInlineKeyboard.MesegeID)
